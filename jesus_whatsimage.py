import time
import pyautogui
from openpyxl import load_workbook

def send_media_file(to='',mediafile=''):
    pyautogui.moveTo(200, 208, duration=1)
    pyautogui.click()
    time.sleep(1)
    pyautogui.typewrite(name)
    time.sleep(3)
    pyautogui.press('enter')
    pyautogui.moveTo(1720, 160, duration=1)
    pyautogui.click()
    pyautogui.moveTo(1720, 250, duration=1)
    pyautogui.click()
    time.sleep(1)
    pyautogui.typewrite(mediafile)
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(1)
    pyautogui.press('enter')
    time.sleep(2)

wb = load_workbook(filename = 'contactExcel.xlsx')
ws = wb.active

contacts = []

#change accordingly
countOfContacts = 3

for row in range(2,countOfContacts+2):
    contacts.append(ws['A'+str(row)].value)

#execution paused for switching current window to whatsApp Web (in browser)
time.sleep(10)
for name in contacts:
    # C:\\ instead of C:\, escape backslash in path
    send_media_file(to=name,mediafile="1.png")
